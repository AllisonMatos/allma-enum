#!/usr/bin/env python3
import sys
import time
from core import runner
from core.colors import C  # Centralizado em core/colors.py


# --------- MÓDULOS ---------
MODULES = {
    "1": "domain",
    "2": "urls",
    "3": "services",
    "4": "files",
    "5": "jsscanner",
    "6": "fingerprint",
    "7": "endpoint",
    "8": "wordlist",
    "9": "sourcemaps",
    "10": "cve",
    "11": "admin",
    "12": "cors",
    "13": "takeover",
    "14": "headers",
    "15": "waf",
    "16": "emails",
    "17": "graphql",
    "18": "jwt_analyzer",
    "19": "api_fuzzer",
    "20": "cloud",
    "21": "host_header_injection",
    "22": "email_security",
    "23": "google_dorks",
    "24": "cookies",
    "25": "asn",
    "26": "screenshots",
    "27": "cache",
    "28": "ssti",
    "30": "ssrf",
    "31": "bypass403",
    "32": "crlf",
    "33": "nuclei",
    "35": "github_dorks",
    "36": "wayback_diff",
    "37": "vhost",
    "38": "cspt",
    "34": "all",
}

# --------- DEPENDÊNCIAS ---------
# Mapeia quais módulos o módulo X depende para rodar completo
DEPENDENCIES = {
    "1": ["1"],
    "2": ["1", "2"],
    "3": ["1", "2", "3"],
    "4": ["1", "2", "3", "4"],
    "5": ["1", "2", "3", "4", "5"],
    "6": ["1", "2", "3", "4", "5", "6"],
    "7": ["1", "2", "3", "4", "5", "6", "7"],
    "8": ["1", "2", "3", "4", "5", "6", "7", "8"],
    "9": ["1", "2", "5", "9"],
    "10": ["1", "2", "3", "6", "10"],
    "11": ["1", "2", "11"],
    "12": ["1", "2", "12"],
    "13": ["1", "2", "13"],
    "14": ["1", "2", "14"],
    "15": ["1", "2", "15"],
    "16": ["1", "2", "16"],
    "17": ["1", "2", "7", "17"],
    "18": ["1", "2", "18"],
    "19": ["1", "2", "7", "19"],
    "20": ["1", "20"],
    "21": ["1", "21"],
    "22": ["1", "22"],
    "23": ["1", "2", "23"],
    "24": ["1", "2", "24"],
    "25": ["1", "25"],
    "26": ["1", "2", "26"],
    "27": ["1", "2", "27"],
    "28": ["1", "2", "28"],
    "30": ["1", "2", "30"],
    "31": ["1", "2", "11", "31"],  # bypass403 depende do admin scanner (encontra 403s)
    "32": ["1", "2", "32"],
    "33": ["1", "2", "33"],
    "35": ["1", "35"],  # github_dorks_active: só precisa do target
    "36": ["1", "2", "5", "36"],  # wayback_diff: depende de jsscanner para js_routes.json
    "37": ["1", "37"],  # vhost: depende de domain para ips.txt
    "38": ["1", "2", "5", "38"],  # cspt: depende de urls (dinâmico) e jsscanner (estático)
    "34": [str(i) for i in range(1, 29)] + ["30", "31", "32", "33", "35", "36", "37", "38"],  # ALL: inclui novos módulos
}


# ---------- BANNER ----------
def print_banner():
    banner = f"""
{C.BOLD}{C.PURPLE}
╔════════════════════════════════════════════════════════════════╗
║{C.CYAN}    ███████╗███╗   ██╗██╗   ██╗███╗   ███╗███████╗██████╗ {C.PURPLE}     ║
║{C.CYAN}    ██╔════╝████╗  ██║██║   ██║████╗ ████║██╔════╝██╔══██╗{C.PURPLE}     ║
║{C.CYAN}    █████╗  ██╔██╗ ██║██║   ██║██╔████╔██║█████╗  ██████╔╝{C.PURPLE}     ║
║{C.CYAN}    ██╔══╝  ██║╚██╗██║██║   ██║██║╚██╔╝██║██╔══╝  ██╔══██╗{C.PURPLE}     ║
║{C.CYAN}    ███████╗██║ ╚████║╚██████╔╝██║ ╚═╝ ██║███████╗██║  ██║{C.PURPLE}     ║
║{C.CYAN}    ╚══════╝╚═╝  ╚═══╝ ╚═════╝ ╚═╝     ╚═╝╚══════╝╚═╝  ╚═╝{C.PURPLE}     ║
║                                                                  ║
║    {C.YELLOW}FERRAMENTA PROFISSIONAL DE ENUMERAÇÃO E RECONHECIMENTO{C.PURPLE}    ║
╚════════════════════════════════════════════════════════════════╝
{C.END}
    """
    print(banner)


# ---------- MENU ----------
def print_menu():
    print(f"\n{C.BOLD}{C.CYAN}📋 MÓDULOS DISPONÍVEIS{C.END}\n")

    modules = {
        "0": ("documentation", "Leia como a ferramenta funciona", "📖"),
        "1": ("domain", "Enumeração de subdomínios e portas", "🌐"),
        "2": ("urls", "Descoberta e validação de URLs", "🔗"),
        "3": ("services", "Identificação de serviços", "🛠️"),
        "4": ("files", "Busca por arquivos sensíveis", "📁"),
        "5": ("jsscanner", "Análise de JavaScript", "⚡"),
        "6": ("fingerprint", "Fingerprinting de tecnologias", "🖐️"),
        "7": ("endpoint", "Enumeração de endpoints API", "🎯"),
        "8": ("wordlist", "Força bruta em diretórios", "🗂️"),
        "9": ("sourcemaps", "Extração e Análise de Source Maps", "🗺️"),
        "10": ("cve", "Varredura de CVEs conhecidas", "🛡️"),
        "11": ("admin", "Busca por painéis administrativos", "🔑"),
        "12": ("cors", "Análise de CORS Config", "🟧"),
        "13": ("takeover", "Subdomain Takeover Check", "🏴‍☠️"),
        "14": ("headers", "Análise de Security Headers", "📜"),
        "15": ("waf", "Detecção de WAF", "🛡️"),
        "16": ("emails", "Extração de e-mails", "📧"),
        "17": ("graphql", "GraphQL Introspection", "🧬"),
        "18": ("jwt_analyzer", "Análise de JWT Tokens", "🔑"),
        "19": ("api_fuzzer", "API Fuzzer (Kiterunner)", "🪁"),
        "20": ("cloud", "Cloud Recon (S3/Azure/GCP)", "🌩️"),
        "21": ("host_header", "Host Header Injection", "🏠"),
        "22": ("email_security", "SPF/DMARC/DKIM Check", "📧"),
        "23": ("google_dorks", "Google Dorks Generator", "🔍"),
        "24": ("cookies", "Análise de Segurança de Cookies", "🍪"),
        "25": ("asn", "CIDR/ASN Mapping", "🌐"),
        "26": ("screenshots", "Screenshot Capture", "📸"),
        "27": ("cache", "Web Cache Vulnerabilities", "⚡"),
        "28": ("ssti", "Server-Side Template Injection", "🧨"),
        "30": ("ssrf", "SSRF Scanner (Active Validation)", "🌐"),
        "31": ("bypass403", "403 Bypass Scanner", "🔓"),
        "32": ("crlf", "CRLF Injection Scanner", "💉"),
        "33": ("nuclei", "Nuclei Vuln Scanner", "☢️"),
        "35": ("github_dorks", "GitHub/GitLab Dorking Ativo", "🐙"),
        "36": ("wayback_diff", "Wayback JS Snapshot Diff", "🕒"),
        "37": ("vhost", "Virtual Host Discovery (ffuf)", "🌐"),
        "38": ("cspt", "Client-Side Path Traversal Scanner", "🔀"),
        "34": ("all", "Execução completa", "🚀")
    }

    for k, (name, desc, emoji) in modules.items():
        print(
            f"  {C.BOLD}{C.YELLOW}{k:>2}{C.END} {C.GREEN}▶{C.END} {emoji} "
            f"{C.BOLD}{name:<12}{C.END} {C.GRAY}─{C.END} {desc}"
        )


# ---------- CONFIG PORTAS (versão antiga) ----------
def ask_ports_mode():
    """Seleção de modo de portas usando opções numéricas."""
    print(f"\n{C.BOLD}{C.CYAN}🌐 CONFIGURAÇÃO DE PORTAS{C.END}")
    print(f"  {C.YELLOW}┌─ Modos disponíveis ───────────────────────────┐{C.END}")
    print(f"  {C.YELLOW}│{C.END} {C.GREEN}1{C.END} → TOP 100 portas                  {C.YELLOW}│{C.END}")
    print(f"  {C.YELLOW}│{C.END} {C.GREEN}2{C.END} → TOP 1000 portas                 {C.YELLOW}│{C.END}")
    print(f"  {C.YELLOW}│{C.END} {C.GREEN}3{C.END} → ALL (todas as portas)            {C.YELLOW}│{C.END}")
    print(f"  {C.YELLOW}└────────────────────────────────────────────────┘{C.END}")

    while True:
        val = input(f"\n{C.BOLD}{C.BLUE}Escolha o modo [3]: {C.END}").strip()

        if val == "":
            return "all"

        if val == "1":
            return "100"

        if val == "2":
            return "1000"

        if val == "3":
            return "all"

        print(f"{C.RED}❌ Opção inválida! Use: 1, 2 ou 3.{C.END}")


# ---------- NMAP ----------
def ask_nmap_args():
    import shlex
    print(f"\n{C.BOLD}{C.CYAN}🔧 CONFIGURAÇÃO DO NMAP{C.END}")
    print(f"{C.YELLOW}Padrão:{C.END} {C.GREEN}-sV -Pn{C.END}")
    val = input(f"{C.BOLD}{C.BLUE}Args: {C.END}").strip()
    raw = val if val else "-sV -Pn"
    # Sanitizar: usar shlex.split para parsing seguro
    try:
        parts = shlex.split(raw)
    except ValueError:
        print(f"{C.RED}❌ Argumentos inválidos, usando padrão -sV -Pn{C.END}")
        return "-sV -Pn"
    # Bloquear flags perigosas
    blocked = {"--script", "--exec", "--interactive", "-iL", "--script-args", "--script-updatedb"}
    for part in parts:
        flag = part.split("=")[0].lower()
        if flag in blocked:
            print(f"{C.RED}❌ Flag bloqueada: {flag}. Usando padrão -sV -Pn{C.END}")
            return "-sV -Pn"
    return raw


# ---------- VALIDAR TARGET ----------
def validate_target(target):
    return bool(target and len(target) >= 3)


# ---------- EXECUTION PLAN ----------
def print_execution_plan(chain, target):
    print(f"\n{C.BOLD}{C.CYAN}📊 PLANO DE EXECUÇÃO{C.END}")
    print(f"{C.YELLOW}Target:{C.END} {C.GREEN}{target}{C.END}")
    print(f"{C.YELLOW}Módulos:{C.END} {C.BLUE}{', '.join(MODULES[c] for c in chain)}{C.END}")
    print(f"{C.YELLOW}Total:{C.END} {len(chain)} fases\n")


# ---------- CARREGAR TARGETS DE ARQUIVO ----------
def load_targets_file(filepath: str) -> list:
    """Carrega targets de um arquivo .txt (1 domínio por linha)."""
    import os
    if not os.path.isfile(filepath):
        print(f"{C.RED}❌ Arquivo não encontrado: {filepath}{C.END}")
        return []
    try:
        with open(filepath, 'r') as f:
            targets = [l.strip() for l in f if l.strip() and not l.strip().startswith("#")]
        # Deduplica mantendo ordem
        seen = set()
        unique = []
        for t in targets:
            if t not in seen:
                seen.add(t)
                unique.append(t)
        return unique
    except Exception as e:
        print(f"{C.RED}❌ Erro ao ler arquivo: {e}{C.END}")
        return []


# ---------- CONFIGURAÇÃO GLOBAL (compartilhada por single/multi) ----------
def ask_global_config():
    """Pergunta configurações globais aplicáveis a todos os targets."""
    config = {}

    # User-Agent
    print(f"\n{C.BOLD}{C.CYAN}🌐 USER-AGENT{C.END}")
    print(f"  {C.YELLOW}Padrão:{C.END} {C.GREEN}Chrome/124 (rotação automática){C.END}")
    print(f"  {C.YELLOW}Tip:{C.END} Alguns programas (Bugcrowd/HackerOne) pedem UA customizado.")
    custom_ua = input(f"{C.BOLD}{C.BLUE}User-Agent customizado (Enter para padrão): {C.END}").strip()
    if custom_ua:
        from core.config import _USER_AGENT_POOL
        import core.config as _cfg
        _cfg.DEFAULT_USER_AGENT = custom_ua
        _cfg._USER_AGENT_POOL = [custom_ua]
        print(f"  {C.GREEN}✅ UA definido: {custom_ua[:60]}{'...' if len(custom_ua) > 60 else ''}{C.END}")
    config["custom_ua"] = custom_ua

    # Proxy
    print(f"\n{C.BOLD}{C.CYAN}🛡️ PROXY (Opcional){C.END}")
    print(f"  {C.YELLOW}Tip:{C.END} Útil para rotear tráfego para Burp Suite ou ZAP.")
    proxy_url = input(f"{C.BOLD}{C.BLUE}Proxy (ex: http://127.0.0.1:8080) [Enter p/ pular]: {C.END}").strip()
    if proxy_url:
        import os
        os.environ["HTTP_PROXY"] = proxy_url
        os.environ["HTTPS_PROXY"] = proxy_url
        os.environ["HTTPX_PROXY"] = proxy_url
        print(f"  {C.GREEN}✅ Todo o tráfego HTTP/HTTPS será roteado para: {proxy_url}{C.END}")
    config["proxy_url"] = proxy_url

    # Deep/Stealth
    print(f"\n{C.BOLD}{C.CYAN}🎯 OPÇÕES DE PERFORMANCE (V10){C.END}")
    config["deep"] = input(f"{C.BOLD}{C.BLUE}Habilitar --deep (varredura profunda)? [s/N]: {C.END}").strip().lower() in ["s", "sim", "y", "yes"]
    config["stealth"] = input(f"{C.BOLD}{C.BLUE}Habilitar --stealth (mais silencioso/lento)? [s/N]: {C.END}").strip().lower() in ["s", "sim", "y", "yes"]

    # Exclusões
    print(f"\n{C.BOLD}{C.CYAN}🚫 EXCLUSÕES (Opcional){C.END}")
    exclude_input = input(f"{C.BOLD}{C.BLUE}Caminho do arquivo (.xlsx/.txt) OU lista de IPs/Hosts (Enter p/ pular): {C.END}").strip()
    exclude_hosts = []

    if exclude_input:
        import os
        if os.path.isfile(exclude_input):
            if exclude_input.endswith('.xlsx'):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(exclude_input, data_only=True)
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        for row in ws.iter_rows(values_only=True):
                            for cell in row:
                                if cell and isinstance(cell, str):
                                    val = str(cell).strip()
                                    if val and " " not in val:
                                        exclude_hosts.append(val)
                except Exception as e:
                    print(f"{C.RED}❌ Erro ao ler XLSX: {e}{C.END}")
            else:
                try:
                    with open(exclude_input, 'r') as f:
                        exclude_hosts = [l.strip() for l in f if l.strip()]
                except Exception as e:
                    print(f"{C.RED}❌ Erro ao ler arquivo: {e}{C.END}")
        else:
            exclude_hosts = [h.strip() for h in exclude_input.split(",") if h.strip()]

        exclude_hosts = list(set(exclude_hosts))
        if exclude_hosts:
            print(f"   {C.YELLOW}⛔ Carregados {len(exclude_hosts)} hosts/IPs para exclusão.{C.END}")

    config["exclude_hosts"] = exclude_hosts
    return config


# ---------- MAIN ----------
def main():
    print("\033c", end="")
    print_banner()
    print_menu()

    print(f"\n{C.BOLD}{C.CYAN}🎯 SELEÇÃO DO MÓDULO{C.END}")
    choice = input(f"\n{C.BOLD}{C.BLUE}Digite o número: {C.END}").strip()

    if choice == "0":
        print(f"\n{C.GREEN}📖 Iniciando Visualizador de Documentação...{C.END}")
        from plugins.documentation import main as docs_main
        docs_main.run_docs()
        sys.exit(0)

    choices = [c.strip() for c in choice.split(",") if c.strip()]
    if not choices:
        print(f"{C.RED}❌ Nenhuma opção fornecida!{C.END}")
        sys.exit(1)

    for c in choices:
        if c not in MODULES:
            print(f"{C.RED}❌ Opção inválida: {c}{C.END}")
            sys.exit(1)

    # ============================================================
    # MODO DE ALVO: Single ou Multi-Target
    # ============================================================
    print(f"\n{C.BOLD}{C.CYAN}🎯 MODO DE ALVO{C.END}")
    print(f"  {C.YELLOW}┌─ Escolha o modo ─────────────────────────────────────┐{C.END}")
    print(f"  {C.YELLOW}│{C.END} {C.GREEN}1{C.END} → Target Único (padrão)                           {C.YELLOW}│{C.END}")
    print(f"  {C.YELLOW}│{C.END} {C.GREEN}2{C.END} → Multi-Target (arquivo .txt, 1 domínio por linha) {C.YELLOW}│{C.END}")
    print(f"  {C.YELLOW}└──────────────────────────────────────────────────────┘{C.END}")

    target_mode = input(f"\n{C.BOLD}{C.BLUE}Modo [1/2]: {C.END}").strip()
    is_multi = (target_mode == "2")

    if is_multi:
        # ============================================================
        # MULTI-TARGET MODE
        # ============================================================
        print(f"\n{C.BOLD}{C.PURPLE}🎯 MULTI-TARGET MODE{C.END}")
        targets_path = input(f"{C.BOLD}{C.BLUE}Caminho do arquivo .txt com targets: {C.END}").strip()
        targets = load_targets_file(targets_path)

        if not targets:
            print(f"{C.RED}❌ Nenhum target válido encontrado no arquivo.{C.END}")
            sys.exit(1)

        print(f"\n{C.BOLD}{C.CYAN}📋 Targets carregados ({len(targets)}):{C.END}")
        for i, t in enumerate(targets, 1):
            print(f"  {C.GREEN}{i:>3}{C.END}. {t}")

        # Config global
        global_config = ask_global_config()

        # Chain
        raw_chain = []
        for c in choices:
            raw_chain.extend(DEPENDENCIES[c] + [c])
        chain = list(dict.fromkeys(raw_chain))
        chain = [c for c in chain if c != "28"]

        # Params (mesma config para todos)
        params = {name: {} for name in MODULES.values()}
        for name in MODULES.values():
            params[name]["scope_root"] = ""  # Setado por target no multi_runner
        params["domain"]["closed_scope"] = []

        if global_config["exclude_hosts"]:
            for name in MODULES.values():
                params[name]["exclude_hosts"] = global_config["exclude_hosts"]

        if "1" in chain:
            params["domain"]["ports"] = ask_ports_mode()

        if "3" in chain:
            params["services"]["nmap_args"] = ask_nmap_args()

        # Plano Multi-Target
        print(f"\n{C.BOLD}{C.CYAN}📊 PLANO DE EXECUÇÃO (MULTI-TARGET){C.END}")
        print(f"{C.YELLOW}Targets:{C.END} {C.GREEN}{len(targets)}{C.END}")
        print(f"{C.YELLOW}Paralelos:{C.END} {C.GREEN}3{C.END}")
        print(f"{C.YELLOW}Módulos:{C.END} {C.BLUE}{', '.join(MODULES[c] for c in chain)}{C.END}")
        print(f"{C.YELLOW}Total:{C.END} {len(chain)} fases por target\n")

        open_terminals_in = input(f"{C.BOLD}{C.BLUE}Abrir terminais popup individuais para acompanhar ao vivo? [s/N]: {C.END}").strip().lower()
        open_terminals = open_terminals_in in ["s", "sim", "y", "yes"]

        confirm = input(f"\n{C.BOLD}{C.BLUE}Iniciar varredura multi-target? [S/n]: {C.END}").strip().lower()

        if confirm in ["", "s", "sim", "y", "yes"]:
            print(f"{C.GREEN}\n🎬 Iniciando Multi-Target...\n{C.END}")
            time.sleep(1)
            from core.multi_runner import execute_multi_target
            execute_multi_target(
                targets=targets,
                chain=chain,
                params=params,
                deep=global_config["deep"],
                stealth=global_config["stealth"],
                max_parallel=3,
                open_terminals=open_terminals,
            )
            print(f"{C.GREEN}{C.BOLD}✔ Multi-Target Concluído!{C.END}")
        else:
            print(f"{C.YELLOW}Operação cancelada.{C.END}")

    else:
        # ============================================================
        # SINGLE TARGET MODE (original)
        # ============================================================
        print(f"\n{C.BOLD}{C.CYAN}🎯 DEFINIÇÃO DO ALVO{C.END}")
        target = input(f"{C.BOLD}{C.BLUE}Domínio/Empresa principal: {C.END}").strip()

        if not validate_target(target):
            print(f"{C.RED}❌ Target inválido.{C.END}")
            sys.exit(1)

        print(f"\n{C.BOLD}{C.CYAN}🌐 DOMÍNIO DE ESCOPO (higiene de URLs){C.END}")
        print(f"{C.GRAY}Se a pasta de output for diferente do domínio do programa (ex.: avenue2.us vs avenue.us),{C.END}")
        print(f"{C.GRAY}defina aqui o domínio raíz usado em is_in_scope. Enter = mesmo que o alvo.{C.END}")
        scope_root_in = input(f"{C.BOLD}{C.BLUE}Domínio raíz para escopo de URLs [{target}]: {C.END}").strip()
        scope_root = scope_root_in or target

        closed_scope_list = []

        print(f"\n{C.BOLD}{C.CYAN}🎯 TIPO DE ESCOPO{C.END}")
        print(f"  {C.YELLOW}┌─ Modo de Descoberta ──────────────────────────┐{C.END}")
        print(f"  {C.YELLOW}│{C.END} {C.GREEN}1{C.END} → Subdomain Discovery Automática (Padrão)   {C.YELLOW}│{C.END}")
        print(f"  {C.YELLOW}│{C.END} {C.GREEN}2{C.END} → Escopo Fechado (Informar subdomínios)     {C.YELLOW}│{C.END}")
        print(f"  {C.YELLOW}└───────────────────────────────────────────────┘{C.END}")

        scope_mode = input(f"\n{C.BOLD}{C.BLUE}Escolha o modo [1/2]: {C.END}").strip() or "1"
        if scope_mode == "2":
            print(f"\n{C.YELLOW}Informe os subdomínios separados por vírgula.{C.END}")
            domains_input = input(f"{C.BOLD}{C.BLUE}Subdomínios: {C.END}").strip()
            if domains_input:
                closed_scope_list = [d.strip() for d in domains_input.split(",") if d.strip()]

            if not closed_scope_list:
                print(f"{C.RED}❌ Nenhum subdomínio fornecido. Usando descoberta padrão...{C.END}")
            else:
                import core.config as _cfg
                _cfg.STRICT_SCOPE_HOSTS = closed_scope_list
                print(f"{C.GREEN}✅ Escopo estrito configurado para {len(closed_scope_list)} host(s).{C.END}")

        # Config global
        global_config = ask_global_config()

        # chain
        raw_chain = []
        for c in choices:
            raw_chain.extend(DEPENDENCIES[c] + [c])
        chain = list(dict.fromkeys(raw_chain))
        chain = [c for c in chain if c != "28"]

        # parâmetros
        params = {name: {} for name in MODULES.values()}
        for name in MODULES.values():
            params[name]["scope_root"] = scope_root

        params["domain"]["closed_scope"] = closed_scope_list

        if global_config["exclude_hosts"]:
            for name in MODULES.values():
                params[name]["exclude_hosts"] = global_config["exclude_hosts"]

        if "1" in chain:
            params["domain"]["ports"] = ask_ports_mode()

        if "3" in chain:
            params["services"]["nmap_args"] = ask_nmap_args()

        # plano
        print_execution_plan(chain, target)

        # iniciar?
        confirm = input(f"{C.BOLD}{C.BLUE}Iniciar varredura? [S/n]: {C.END}").strip().lower()

        if confirm in ["", "s", "sim", "y", "yes"]:
            print(f"{C.GREEN}\n🎬 Iniciando...\n{C.END}")
            time.sleep(1)
            runner.execute_chain(target, chain, params, deep=global_config["deep"], stealth=global_config["stealth"])
            print(f"{C.GREEN}{C.BOLD}✔ Concluído!{C.END}")
        else:
            print(f"{C.YELLOW}Operação cancelada.{C.END}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print(f"\n{C.RED}Interrompido pelo usuário.{C.END}")
        sys.exit(1)
    except Exception as e:
        print(f"\n{C.RED}Erro inesperado: {e}{C.END}")
        sys.exit(1)
